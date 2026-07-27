"""
movik/03_build_master.py
========================
Excel maestro plano: cruza los carriers de movik_carriers.xlsx con los
resultados UCC de movik.db y exporta UNA sola hoja "Maestro".

Lee:    movik_carriers.xlsx  (hojas FL, TX, CA)  ← 02_transform_export.py
        movik.db             (ucc_filings, scrape_log)  ← ucc_scraper.py
Genera: movik_maestro.xlsx   (hoja única "Maestro")

Uso:
  python 03_build_master.py

REGLA CLAVE — ausencia de dato ≠ ausencia de UCC
------------------------------------------------
Un carrier solo se marca P1 ("sin UCC, llamar ya") si REALMENTE se consultó
el registro y no hubo match (scrape_log.status = 'not_found').

  - Nunca consultado (TX, CA, o FL aún pendiente) → "Sin datos", NO P1.
  - Consulta con error (timeout, 429)             → "Error", NO P1.

Marcar como P1 a quien nunca se consultó convertiría 500k+ carriers sin
scrapear en leads falsos indistinguibles de los legítimos.

Prioridades (ver ucc_scraper.ScrapeResult.alert_priority):
  P1   sin UCC registrado
  P2a  status Lapsed (venció, sin fecha exacta)   ← lo que da el API de FL hoy
  P2b  Filed + expires_date <= 30 días           ← requiere bulk data
  P2c  Filed + expires_date 31-60 días           ← requiere bulk data
  P3   Filed + expires_date 61-365 días          ← requiere bulk data
  P4   Filed activo (sin fecha o > 365 días)
"""

import sys
import sqlite3
from pathlib import Path
from collections import Counter

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

try:
    import duckdb
    import openpyxl
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Falta: {e}  →  pip install duckdb openpyxl")
    sys.exit(1)

# ── CONFIG ───────────────────────────────────────────────────────────────────
CARRIERS_XLSX = "movik_carriers.xlsx"
DB_FILE       = "movik.db"
OUTPUT_XLSX   = "movik_maestro.xlsx"

SCRAPED_STATES = {"FL"}   # estados con scraper implementado y corrido

# Orden exacto de columnas del maestro
COLUMNS = [
    "alert_priority", "phy_state", "dot_number", "legal_name", "dba_name",
    "classdef", "carrier_operation", "phone", "cell_phone", "email_address",
    "company_officer_1", "phy_city", "phy_street", "phy_zip", "power_units",
    "truck_units", "fleetsize", "total_drivers", "safety_rating", "mcs150_date",
    "ucc_found", "ucc_number", "date_filed", "expires_date",
    "secured_party", "days_to_expiry", "notes",
]

# Header amigable de movik_carriers.xlsx → nombre de columna del maestro
HEADER_MAP = {
    "DOT #": "dot_number",
    "Legal Name": "legal_name",
    "DBA": "dba_name",
    "Class": "classdef",
    "Operation": "carrier_operation",
    "State": "phy_state",
    "City": "phy_city",
    "Street": "phy_street",
    "ZIP": "phy_zip",
    "Phone": "phone",
    "Cell": "cell_phone",
    "Email": "email_address",
    "Officer 1": "company_officer_1",
    "Power Units": "power_units",
    "Trucks": "truck_units",
    "Fleet Size": "fleetsize",
    "Drivers": "total_drivers",
    "Safety Rating": "safety_rating",
    "MCS-150 Date": "mcs150_date",
}

# Orden de urgencia para el sort (alert_priority ASC = más urgente primero)
PRIORITY_RANK = {
    "P1": 0, "P2a": 1, "P2b": 2, "P2c": 3, "P3": 4, "P4": 5,
    "Error": 6, "Sin datos": 7,
}

# Color de fila por prioridad. P2a/P2b/P2c comparten el color de P2.
ROW_FILLS = {
    "P1":        "FFF0F0",   # rojo muy claro
    "P2":        "FFF7ED",   # naranja muy claro
    "P3":        "FEFCE8",   # amarillo muy claro
    "P4":        "FFFFFF",   # blanco
    "Sin datos": "F5F5F5",   # gris muy claro
    "Error":     "F5F5F5",
}

COL_WIDTHS = {
    "alert_priority": 13, "phy_state": 7, "dot_number": 10, "legal_name": 32,
    "dba_name": 20, "phone": 15, "email_address": 26, "phy_city": 16,
    "secured_party": 30, "days_to_expiry": 12, "notes": 20,
}

HEADER_BG   = "1D5FBD"
HEADER_FG   = "FFFFFF"


def fill_for(priority: str) -> str:
    if priority.startswith("P2"):
        return ROW_FILLS["P2"]
    return ROW_FILLS.get(priority, "FFFFFF")


# ── UCC DATA ─────────────────────────────────────────────────────────────────
def load_ucc(db_path: str) -> dict:
    """
    {dot_number: {...}} para carriers que SÍ se consultaron.
    Los que no aparecen aquí nunca se consultaron → "Sin datos".
    """
    if not Path(db_path).exists():
        print(f"⚠️  {db_path} no existe: ningún carrier tiene datos UCC todavía.")
        return {}

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    # scrape_log es la fuente de verdad de "se consultó y con qué resultado".
    log = {}
    for r in conn.execute("SELECT dot_number, status, error_msg FROM scrape_log"):
        log[str(r["dot_number"])] = {"status": r["status"], "error": r["error_msg"]}

    # Filings agrupados por carrier.
    filings = {}
    for r in conn.execute("""
        SELECT dot_number, ucc_number, date_filed, expires_date, days_to_expiry,
               secured_party, filing_status, alert_priority
        FROM ucc_filings WHERE match_found = 1
    """):
        filings.setdefault(str(r["dot_number"]), []).append(dict(r))
    conn.close()

    out = {}
    for dot, entry in log.items():
        st = entry["status"]

        if st == "error":
            out[dot] = {
                "alert_priority": "Error", "ucc_found": "Error",
                "ucc_number": "", "date_filed": "", "expires_date": "",
                "secured_party": "", "days_to_expiry": None,
                "notes": (entry["error"] or "")[:200],
            }
            continue

        if st == "not_found":
            out[dot] = {
                "alert_priority": "P1", "ucc_found": "No",
                "ucc_number": "", "date_filed": "", "expires_date": "",
                "secured_party": "", "days_to_expiry": None, "notes": "",
            }
            continue

        fs = filings.get(dot) or []
        if not fs:
            # 'found' sin filings no debería pasar; si pasa, no inventamos P1.
            out[dot] = {
                "alert_priority": "Error", "ucc_found": "Error",
                "ucc_number": "", "date_filed": "", "expires_date": "",
                "secured_party": "", "days_to_expiry": None,
                "notes": "found sin filings en ucc_filings",
            }
            continue

        # Representativo: el gravamen activo manda; si no hay, el primero.
        active = [f for f in fs if (f["filing_status"] or "").lower() == "filed"]
        rep = (active or fs)[0]

        n_filed  = len(active)
        n_lapsed = sum(1 for f in fs if (f["filing_status"] or "").lower() == "lapsed")
        note = ""
        if len(fs) > 1:
            note = f"{len(fs)} filings ({n_filed} filed, {n_lapsed} lapsed)"

        out[dot] = {
            "alert_priority": rep["alert_priority"] or "P4",
            "ucc_found": "Sí",
            "ucc_number": rep["ucc_number"] or "",
            "date_filed": rep["date_filed"] or "",
            "expires_date": rep["expires_date"] or "",
            "secured_party": rep["secured_party"] or "",
            "days_to_expiry": rep["days_to_expiry"],
            "notes": note,
        }

    return out


# ── LECTURA DE CARRIERS ──────────────────────────────────────────────────────
def iter_carriers(xlsx_path: str):
    """Stream de filas de todas las hojas de estado (read_only: RAM plana)."""
    if not Path(xlsx_path).exists():
        print(f"❌ {xlsx_path} no existe. Corre 02_transform_export.py primero.")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        if sheet == "Resumen":
            continue
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration:
            continue
        idx = {h: i for i, h in enumerate(headers) if h in HEADER_MAP}
        for row in rows:
            rec = {HEADER_MAP[h]: (row[i] if row[i] is not None else "")
                   for h, i in idx.items()}
            rec.setdefault("phy_state", sheet)
            if not rec.get("phy_state"):
                rec["phy_state"] = sheet
            yield rec
        print(f"  · hoja {sheet} leída")
    wb.close()


def build_rows(ucc: dict):
    """Genera las filas del maestro (dicts en el orden de COLUMNS)."""
    for rec in iter_carriers(CARRIERS_XLSX):
        dot   = str(rec.get("dot_number") or "").strip()
        state = (rec.get("phy_state") or "").strip()

        u = ucc.get(dot)
        if u is None:
            # Nunca consultado. NO es P1.
            u = {
                "alert_priority": "Sin datos",
                "ucc_found": f"Sin datos ({state})" if state else "Sin datos",
                "ucc_number": "", "date_filed": "", "expires_date": "",
                "secured_party": "", "days_to_expiry": None,
                "notes": "" if state in SCRAPED_STATES else f"{state} sin scraper UCC",
            }

        out = {c: rec.get(c, "") for c in COLUMNS}
        out.update(u)
        out["dot_number"] = dot
        out["phy_state"]  = state
        yield out


# ── EXCEL ────────────────────────────────────────────────────────────────────
def write_master(con, total: int, out_path: str):
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(title="Maestro")

    header_font = Font(name="Arial", bold=True, color=HEADER_FG, size=10)
    header_fill = PatternFill("solid", fgColor=HEADER_BG)
    header_algn = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font   = Font(name="Arial", size=9)
    fills       = {k: PatternFill("solid", fgColor=v) for k, v in ROW_FILLS.items()}

    for i, c in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = COL_WIDTHS.get(c, 12)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    ws.row_dimensions[1].height = 28

    hdr = []
    for c in COLUMNS:
        cell = WriteOnlyCell(ws, value=c)
        cell.font, cell.fill, cell.alignment = header_font, header_fill, header_algn
        hdr.append(cell)
    ws.append(hdr)

    counts_state = Counter()
    counts_prio  = Counter()
    written = 0

    cur = con.execute(f"""
        SELECT {', '.join(COLUMNS)}
        FROM maestro
        ORDER BY prio_rank ASC, phy_state ASC, legal_name ASC
    """)

    while True:
        chunk = cur.fetchmany(20_000)
        if not chunk:
            break
        for row in chunk:
            rec  = dict(zip(COLUMNS, row))
            prio = rec["alert_priority"] or "Sin datos"
            key  = "P2" if prio.startswith("P2") else prio
            fill = fills.get(key, fills["P4"])

            cells = []
            for c in COLUMNS:
                v = rec[c]
                cell = WriteOnlyCell(ws, value=("" if v is None else v))
                cell.font = data_font
                cell.fill = fill
                cells.append(cell)
            ws.append(cells)

            counts_state[rec["phy_state"]] += 1
            counts_prio[prio] += 1
            written += 1

        print(f"  · {written:,}/{total:,} filas escritas")

    wb.save(out_path)
    return written, counts_state, counts_prio


# ── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    print("\n📊 Movik — Build Master")
    print(f"   Carriers: {CARRIERS_XLSX}")
    print(f"   UCC:      {DB_FILE}")
    print(f"   Destino:  {OUTPUT_XLSX}\n")

    ucc = load_ucc(DB_FILE)
    print(f"✅ {len(ucc):,} carriers con resultado de scrape en {DB_FILE}")

    # DuckDB hace el join/orden en disco: 668k filas no caben cómodas en RAM.
    con = duckdb.connect()
    cols_ddl = ", ".join(f'"{c}" VARCHAR' for c in COLUMNS if c != "days_to_expiry")
    con.execute(f"""
        CREATE TABLE maestro (
            {cols_ddl},
            "days_to_expiry" INTEGER,
            prio_rank INTEGER
        )
    """)

    insert_cols = COLUMNS + ["prio_rank"]
    placeholders = ",".join("?" * len(insert_cols))
    sql = f'INSERT INTO maestro ({", ".join(chr(34)+c+chr(34) for c in insert_cols)}) VALUES ({placeholders})'

    print("\n📥 Leyendo carriers y cruzando con UCC...")
    batch, total = [], 0
    for row in build_rows(ucc):
        prio = row["alert_priority"] or "Sin datos"
        vals = []
        for c in COLUMNS:
            v = row.get(c)
            if c == "days_to_expiry":
                vals.append(int(v) if v not in (None, "") else None)
            else:
                vals.append("" if v is None else str(v))
        vals.append(PRIORITY_RANK.get(prio, 9))
        batch.append(vals)
        total += 1
        if len(batch) >= 50_000:
            con.executemany(sql, batch)
            batch.clear()
            print(f"  · {total:,} filas cargadas")
    if batch:
        con.executemany(sql, batch)
    print(f"✅ {total:,} filas cargadas en total")

    print("\n📤 Escribiendo Excel...")
    written, counts_state, counts_prio = write_master(con, total, OUTPUT_XLSX)

    # ── Reporte ──────────────────────────────────────────────────────────
    print(f"\n💾 {OUTPUT_XLSX}")
    print(f"\n{'='*46}")
    print(f"TOTAL DE FILAS: {written:,}")

    print("\nPOR ESTADO:")
    for s in sorted(counts_state):
        print(f"  {s:<12} {counts_state[s]:>9,}")

    print("\nPOR PRIORIDAD:")
    order = ["P1", "P2a", "P2b", "P2c", "P3", "P4", "Error", "Sin datos"]
    for p in order:
        if counts_prio.get(p):
            print(f"  {p:<12} {counts_prio[p]:>9,}")
    for p in sorted(set(counts_prio) - set(order)):
        print(f"  {p:<12} {counts_prio[p]:>9,}")

    errors = counts_prio.get("Error", 0)
    print(f"\nERRORES: {errors:,}")
    print(f"{'='*46}")


if __name__ == "__main__":
    main()
