"""
movik/02_transform_export.py
============================
Limpieza, transformación y export a Excel de carriers FL + TX + CA.

Lee: census_file_full.csv  (generado por 01_census_incremental.py)
Genera: movik_carriers.xlsx  (listo para ver y para enriquecer con UCC)

Columnas que produce:
  — Identidad:  dot_number, legal_name, dba_name, status, org_type, classdef
  — Contacto:   phone, cell_phone, email, officer_1, officer_2
  — Dirección:  state, city, street, zip
  — Flota:      power_units, truck_units, fleetsize, total_drivers, total_cdl
  — Actividad:  mcs150_date, add_date, safety_rating, carrier_operation, hm_ind
  — UCC (vacío para el scraper): ucc_found, ucc_number, date_filed,
                                  expires_date, secured_party, days_to_expiry,
                                  alert_priority, notes

Uso:
  python 02_transform_export.py
  python 02_transform_export.py --states FL CA   # solo esos estados
  python 02_transform_export.py --limit 5000     # muestra más pequeña
"""

import argparse
import sys
from pathlib import Path
from datetime import datetime

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

try:
    import duckdb
    import openpyxl
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Falta: {e}  →  pip install duckdb openpyxl")
    sys.exit(1)

# ── CONFIG ───────────────────────────────────────────────────────────────────
CSV_FILE    = "census_file_full.csv"
OUTPUT_XLSX = "movik_carriers.xlsx"
STATES      = ["FL", "TX", "CA"]

# Columnas FMCSA que conservamos (en orden para el Excel)
FMCSA_COLS = [
    ("dot_number",        "DOT #"),
    ("legal_name",        "Legal Name"),
    ("dba_name",          "DBA"),
    ("status_code",       "Status"),
    ("business_org_desc", "Org Type"),
    ("classdef",          "Class"),
    ("carrier_operation", "Operation"),
    ("hm_ind",            "HazMat"),
    ("phy_state",         "State"),
    ("phy_city",          "City"),
    ("phy_street",        "Street"),
    ("phy_zip",           "ZIP"),
    ("phone",             "Phone"),
    ("cell_phone",        "Cell"),
    ("email_address",     "Email"),
    ("company_officer_1", "Officer 1"),
    ("company_officer_2", "Officer 2"),
    ("power_units",       "Power Units"),
    ("truck_units",       "Trucks"),
    ("fleetsize",         "Fleet Size"),
    ("total_drivers",     "Drivers"),
    ("total_cdl",         "CDL Drivers"),
    ("safety_rating",     "Safety Rating"),
    ("mcs150_date",       "MCS-150 Date"),
    ("add_date",          "Add Date"),
]

# Columnas vacías para el scraper UCC (se llenarán después)
UCC_COLS = [
    "ucc_found",       # Sí / No / Pendiente
    "ucc_number",      # 202002233991
    "date_filed",      # 06/12/2020
    "expires_date",    # 06/12/2030
    "secured_party",   # U.S. Small Business Administration
    "days_to_expiry",  # 547
    "alert_priority",  # P1 / P2 / P3 / P4
    "notes",           # notas manuales del equipo
]

# ── COLORES POR ESTADO (header row de cada hoja) ──────────────────────────
STATE_COLORS = {
    "FL": "1D5FBD",   # azul Movik
    "TX": "1A4A8A",   # azul oscuro
    "CA": "153566",   # azul navy
}

HEADER_FONT_COLOR = "FFFFFF"
ALT_ROW_FILL      = "EEF4FF"   # azul muy claro para filas alternadas


# ── HELPERS ──────────────────────────────────────────────────────────────────
def fmt_phone(p: str | None) -> str:
    if not p:
        return ""
    digits = "".join(c for c in str(p) if c.isdigit())
    if len(digits) == 10:
        return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    if len(digits) == 11 and digits[0] == "1":
        return f"({digits[1:4]}) {digits[4:7]}-{digits[7:]}"
    return p.strip()


def fmt_date(d: str | None) -> str:
    """YYYYMMDD → MM/DD/YYYY"""
    if not d:
        return ""
    d = str(d).strip()
    if len(d) == 8 and d.isdigit():
        return f"{d[4:6]}/{d[6:8]}/{d[:4]}"
    return d


def normalize_name(n: str | None) -> str:
    if not n:
        return ""
    return " ".join(str(n).strip().upper().split())


# ── DATA LOADING ──────────────────────────────────────────────────────────────
def load_carriers(csv_path: str, states: list[str], limit: int | None = None) -> dict[str, list]:
    """
    Usa DuckDB para filtrar el CSV directamente sin cargar en RAM.
    Retorna dict {state: [list of dicts]}
    """
    if not Path(csv_path).exists():
        print(f"❌ {csv_path} no encontrado. Corre 01_census_incremental.py primero.")
        sys.exit(1)

    fmcsa_db_cols = [col for col, _ in FMCSA_COLS]
    select_cols   = ", ".join(fmcsa_db_cols)
    st_list       = ", ".join(f"'{s}'" for s in states)
    lim           = f"LIMIT {limit}" if limit else ""

    query = f"""
        SELECT {select_cols}
        FROM read_csv_auto(
            '{csv_path}',
            all_varchar = true,
            header      = true,
            ignore_errors = true
        )
        WHERE phy_state IN ({st_list})
          AND status_code = 'A'
          AND legal_name IS NOT NULL
          AND trim(legal_name) <> ''
        ORDER BY phy_state, mcs150_date DESC NULLS LAST
        {lim}
    """

    con   = duckdb.connect()
    rows  = con.execute(query).fetchall()
    cols  = fmcsa_db_cols

    # Agrupar por estado
    result: dict[str, list] = {s: [] for s in states}
    for row in rows:
        d     = dict(zip(cols, row))
        state = d.get("phy_state", "")
        if state in result:
            result[state].append(d)

    for s in states:
        print(f"  {s}: {len(result[s]):,} carriers activos")

    return result


# ── EXCEL WRITER ──────────────────────────────────────────────────────────────
def write_excel(data: dict[str, list], output_path: str):
    # write_only: las filas se serializan a disco al hacer append en vez de
    # quedarse en RAM. Con 668k filas x 33 columnas, el modo normal construye
    # ~22M objetos Cell y muere por memoria.
    wb = openpyxl.Workbook(write_only=True)

    all_headers = [h for _, h in FMCSA_COLS] + UCC_COLS

    # Estilos compartidos: se crean UNA vez y se reusan en todas las celdas.
    header_font = Font(name="Arial", bold=True, color=HEADER_FONT_COLOR, size=10)
    header_algn = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ucc_fill    = PatternFill("solid", fgColor="0F6E56")   # verde oscuro para UCC
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, bottom=thin)
    alt_fill    = PatternFill("solid", fgColor=ALT_ROW_FILL)
    data_font   = Font(name="Arial", size=9)
    ucc_pending_fill = PatternFill("solid", fgColor="FFF9C4")  # amarillo pálido

    # ── Hoja de resumen ───────────────────────────────────────────────────
    # En write_only el orden de las hojas es el de creación, así que Resumen
    # va primero. `data` ya está materializado, así que los totales se saben.
    total = sum(len(v) for v in data.values())
    ws_sum = wb.create_sheet(title="Resumen")
    for col in ["A", "B", "C"]:
        ws_sum.column_dimensions[col].width = 20

    def _srow(ws, values, font=None):
        cells = []
        for v in values:
            c = WriteOnlyCell(ws, value=v)
            if font:
                c.font = font
            cells.append(c)
        ws.append(cells)

    bold = Font(name="Arial", bold=True)
    _srow(ws_sum, ["Movik — Resumen de carriers"],
          Font(name="Arial", bold=True, size=14, color="1D5FBD"))
    _srow(ws_sum, [])
    _srow(ws_sum, ["Estado", "Carriers", "% del total"], bold)
    for state, carriers in data.items():
        _srow(ws_sum, [state, len(carriers), f"{len(carriers)/max(total,1)*100:.1f}%"])
    _srow(ws_sum, ["TOTAL", total], bold)
    _srow(ws_sum, [f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"],
          Font(name="Arial", size=9, color="888888"))
    _srow(ws_sum, ["Columnas verdes oscuro = vacías, se llenarán con el scraper UCC"],
          Font(name="Arial", size=9, color="0F6E56", italic=True))
    _srow(ws_sum, ["P1=sin UCC (llamar ya) | P2=vence ≤30d | P3=31-60d | P4=activo"],
          Font(name="Arial", size=9, color="888888", italic=True))

    for state, carriers in data.items():
        if not carriers:
            continue

        ws = wb.create_sheet(title=state)
        color = STATE_COLORS.get(state, "1D5FBD")
        header_fill = PatternFill("solid", fgColor=color)

        # ── Anchos, freeze y filtro: antes de escribir filas ───────────────
        col_widths = {
            "DOT #": 10, "Legal Name": 30, "DBA": 20, "Status": 8,
            "Org Type": 14, "Class": 18, "Operation": 10, "HazMat": 8,
            "State": 6, "City": 16, "Street": 24, "ZIP": 8,
            "Phone": 15, "Cell": 15, "Email": 24,
            "Officer 1": 20, "Officer 2": 20,
            "Power Units": 10, "Trucks": 8, "Fleet Size": 9,
            "Drivers": 8, "CDL Drivers": 10, "Safety Rating": 12,
            "MCS-150 Date": 12, "Add Date": 10,
            "ucc_found": 10, "ucc_number": 16, "date_filed": 10,
            "expires_date": 10, "secured_party": 28, "days_to_expiry": 12,
            "alert_priority": 10, "notes": 20,
        }
        for col_idx, header in enumerate(all_headers, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(header, 12)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(all_headers))}1"
        ws.row_dimensions[1].height = 30

        # ── Header row ────────────────────────────────────────────────────
        header_cells = []
        for header in all_headers:
            c = WriteOnlyCell(ws, value=header)
            c.fill      = ucc_fill if header in UCC_COLS else header_fill
            c.font      = header_font
            c.alignment = header_algn
            c.border    = border
            header_cells.append(c)
        ws.append(header_cells)

        # ── Data rows ─────────────────────────────────────────────────────
        for row_idx, carrier in enumerate(carriers, 2):
            fill = alt_fill if row_idx % 2 == 0 else None
            row = []

            for db_col, _ in FMCSA_COLS:
                raw = carrier.get(db_col) or ""

                # Transformaciones de limpieza
                if db_col in ("legal_name", "dba_name"):
                    val = normalize_name(raw)
                elif db_col in ("phone", "cell_phone"):
                    val = fmt_phone(raw)
                elif db_col in ("mcs150_date", "add_date"):
                    val = fmt_date(raw)
                elif db_col == "hm_ind":
                    val = "Sí" if str(raw).upper() == "Y" else ("No" if raw else "")
                elif db_col == "status_code":
                    val = "Activo" if str(raw).upper() == "A" else raw
                else:
                    val = str(raw).strip() if raw else ""

                c = WriteOnlyCell(ws, value=val)
                c.font   = data_font
                c.border = border
                if fill:
                    c.fill = fill
                row.append(c)

            # Columnas UCC vacías con fill amarillo pálido (indican "pendiente")
            for _ in UCC_COLS:
                c = WriteOnlyCell(ws, value="")
                c.fill   = ucc_pending_fill
                c.font   = data_font
                c.border = border
                row.append(c)

            ws.append(row)

        print(f"  ✅ Hoja {state}: {len(carriers):,} filas escritas")

    wb.save(output_path)
    print(f"\n💾 Excel guardado: {output_path}")
    print(f"   Total: {total:,} carriers en {len(data)} estados")


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Movik — Transform + Export Excel")
    parser.add_argument("--states", nargs="+", default=STATES, metavar="ST")
    parser.add_argument("--limit",  type=int,  default=None,  metavar="N",
                        help="Limitar filas por estado (para pruebas)")
    parser.add_argument("--output", type=str,  default=OUTPUT_XLSX)
    args = parser.parse_args()

    print(f"\n📊 Movik Transform — {args.states}")
    print(f"   Fuente:  {CSV_FILE}")
    print(f"   Destino: {args.output}\n")

    data = load_carriers(CSV_FILE, args.states, limit=args.limit)

    if all(len(v) == 0 for v in data.values()):
        print("❌ Sin datos. Verifica que census_file_full.csv existe y tiene datos.")
        sys.exit(1)

    write_excel(data, args.output)
    print(f"\n✅ Abre {args.output} en Excel para revisar.")
    print("   Las columnas amarillas (UCC) se llenarán con ucc_scraper.py")


if __name__ == "__main__":
    main()